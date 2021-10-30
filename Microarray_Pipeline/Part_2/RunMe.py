###################################################################################

import time
from os import listdir
import pandas as pd
import numpy as np
from pandas import ExcelWriter
from openpyxl import load_workbook

###################################################################################

filesystem = pd.read_excel(open('Hierarchy.xlsx','rb'))

t1 = time.time()

###################################################################################

# for index, row in filesystem.iterrows():
# 	platformpathpath 	= 	'DEG_Identification/'+row['Accession']+'/Platform_info/'+row['PlatformInfo']
# 	filepath 		 	= 	'DEG_Identification/'+row['Accession']+'/'+row['FileName']
	
# 	if row['Control'] == 'Yes':
# 		sheet = 'ABS_CALL'
# 	else:
# 		if row['Chip']!= 'Nil':
# 			sheet = 'COMMN_'+row['Chip']+str(row['TimePoint'])+'min'
		
# 		else:
# 			sheet = 'COMMN_'+str(row['TimePoint'])+'min'

# 	print "Converting Probe2Gene for File : "+row['FileName']
	
# 	platforminfo 		= 	pd.read_excel(open(platformpathpath,'rb')).iloc[:,:2]
# 	lastsheet	 		= 	pd.read_excel(open(filepath,'rb'), sheet_name=sheet)
# 	mapped 				= 	pd.merge(platforminfo, lastsheet, on='Probename')

# 	writernw 			= 	ExcelWriter(filepath, engine='openpyxl')
# 	book 				= 	load_workbook(filepath)
# 	writernw.book 		= 	book
# 	writernw.sheets 	= 	dict((ws.title, ws) for ws in book.worksheets)    

# 	mapped.to_excel(writernw, 'Probe2Gene', index=False)
# 	writernw.save()

###################################################################################

filesystem = filesystem[filesystem['Control']=='No']

###################################################################################

# for index, row in filesystem.iterrows():
# 	filepath 		 	= 	'DEG_Identification/'+row['Accession']+'/'+row['FileName']
# 	controlpath 	 	= 	'DEG_Identification/'+row['Accession']+'/'+row['ControlFile']
	
# 	print "Adding Control Values for File : "+row['FileName']
	
# 	lastsheet	 		= 	pd.read_excel(open(filepath,'rb'), sheet_name='Probe2Gene')
# 	controlsheet 		= 	pd.read_excel(open(controlpath,'rb'), sheet_name='Probe2Gene', usecols = 'A,E')

# 	controlinfo			= 	pd.merge(lastsheet, controlsheet, on='Probename')

# 	writernw 			= 	ExcelWriter(filepath, engine='openpyxl')
# 	book 				= 	load_workbook(filepath)
# 	writernw.book 		= 	book
# 	writernw.sheets 	= 	dict((ws.title, ws) for ws in book.worksheets)    

# 	controlinfo.to_excel(writernw, 'T_Vs_CNTRL', index=False)
# 	writernw.save()

###################################################################################


# for index, row in filesystem.iterrows():

# 	print 'Calculating Fold Change for file : '+row['FileName']

# 	filepath 		= 		'DEG_Identification/'+row['Accession']+'/'+row['FileName']
# 	data			= 		pd.read_excel(open(filepath ,'rb'), sheet_name='T_Vs_CNTRL')
# 	valueheader		=		'VALUE_'+str(row['TimePoint'])+'min'
# 	columns 		= 		['Gene Symbol',valueheader,	'VALUE_Control', 'Ratio', 'LFC']
# 	result 			= 		pd.DataFrame(columns = columns)
# 	genes 			= 		list(set(data['Gene Symbol'].values))
	
# 	for gene in genes:
# 		sec = data[data['Gene Symbol']==gene]
# 		repeat = sec.shape[0]
# 		if repeat>1:
# 			Vc = np.mean(sec['VALUE_control'])
# 			Vt = np.mean(sec[valueheader])
# 			Ratio = Vt/Vc
# 			Fc = np.log2(Ratio)
# 			value = [sec['Gene Symbol'].values[0], Vt, Vc, Ratio, Fc]
# 			result.loc[len(result)] =  value
# 		elif repeat==1:
# 			Vc = sec['VALUE_control'].values[0]
# 			Vt = sec[valueheader].values[0]
# 			Ratio = float(Vt)/Vc
# 			Fc = np.log2(Ratio)
# 			value = [sec['Gene Symbol'].values[0], Vt, Vc, Ratio, Fc]
# 			result.loc[len(result)] =  value

# 	writernw 			= 	ExcelWriter(filepath, engine='openpyxl')
# 	book 				= 	load_workbook(filepath)
# 	writernw.book 		= 	book
# 	writernw.sheets 	= 	dict((ws.title, ws) for ws in book.worksheets)    

# 	result.to_excel(writernw, 'AVG_FC_CALC', index=False)
# 	writernw.save()

###################################################################################

for index, row in filesystem.iterrows():
	
	filepath 		 	= 	'DEG_Identification/'+row['Accession']+'/'+row['FileName']
	
	print "Finding Up & Down Regulated for File : "+row['FileName']
	
	lastsheet	 		= 	pd.read_excel(open(filepath,'rb'), sheet_name='AVG_FC_CALC')
	UpRegulatedGenes	=	lastsheet[lastsheet['LFC']>=1]
	DownRegulatedGenes	=	lastsheet[lastsheet['LFC']<=-1]

	writernw 			= 	ExcelWriter(filepath, engine='openpyxl')
	book 				= 	load_workbook(filepath)
	writernw.book 		= 	book
	writernw.sheets 	= 	dict((ws.title, ws) for ws in book.worksheets)    

	UpRegulatedGenes.to_excel(writernw, 'UpRegulated', index=False)
	writernw.save()

	writernw 			= 	ExcelWriter(filepath, engine='openpyxl')
	book 				= 	load_workbook(filepath)
	writernw.book 		= 	book
	writernw.sheets 	= 	dict((ws.title, ws) for ws in book.worksheets)    

	DownRegulatedGenes.to_excel(writernw, 'DownRegulated', index=False)
	writernw.save()


###################################################################################

t2 = time.time()

m, s = divmod(t2-t1, 60)
h, m = divmod(m, 60)
d, h = divmod(h, 24)

print "Done Computation in %d Days, %d Hours, %02d Minutes, %02d Seconds" % (d, h, m, s)

###################################################################################