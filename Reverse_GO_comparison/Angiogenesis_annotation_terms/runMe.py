import pandas as pd
import numpy as np
from pandas import ExcelWriter
from openpyxl import load_workbook

files 		= 	pd.read_excel(open('list.xlsx','rb'), sheet_name=0, index=False)['files']


for file in files:
	accession 	=	file.split('Annot_')[1].split('.')[0]

	data  		= 	pd.read_excel(open(file,'rb'), sheet_name='Total_DEGs', index=False).values
	data 		= 	data[~pd.isnull(data)]

	genesTotal 	= 	list(set(data))
	result	 	= 	pd.DataFrame()
	result['Unique_Genes'] = genesTotal

	writer = ExcelWriter('Result/'+accession+'.xlsx')
	result.to_excel(writer,'Total_DEGs', index =False)
	writer.save()

	datadown  				= 	pd.read_excel(open(file,'rb'), sheet_name='DownRegulated', index=False).values
	datadown 				= 	datadown[~pd.isnull(datadown)]

	genesDown				= 	list(set(datadown))
	downregulated			= 	pd.DataFrame()
	downregulated['Unique_Genes'] 	= 	genesDown

	writernw 			= 	ExcelWriter('Result/'+accession+'.xlsx', engine='openpyxl')
	book 				= 	load_workbook('Result/'+accession+'.xlsx')
	writernw.book 		= 	book
	writernw.sheets 	= 	dict((ws.title, ws) for ws in book.worksheets)    

	downregulated.to_excel(writernw, 'DownRegulated', index=False)
	writernw.save()

	print file
	dataUp					= 	pd.read_excel(open(file,'rb'), sheet_name='UpRegulated', index=False).values
	dataUp 					= 	dataUp[~pd.isnull(dataUp)]

	genesUp					= 	list(set(dataUp))
	upregulated	 			= 	pd.DataFrame()
	upregulated['Unique_Genes'] 	= 	genesUp

	

	writernw 			= 	ExcelWriter('Result/'+accession+'.xlsx', engine='openpyxl')
	book 				= 	load_workbook('Result/'+accession+'.xlsx')
	writernw.book 		= 	book
	writernw.sheets 	= 	dict((ws.title, ws) for ws in book.worksheets)    

	upregulated.to_excel(writernw, 'UpRegulated', index=False)
	writernw.save()



	
